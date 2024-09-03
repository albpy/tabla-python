import cv2 as cv
from utils import *
from constants import *
from math import sqrt
import pytesseract
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class Image:
    def __init__(self, image:cv.Mat):
        self.original_img = image

    def process_img(self):
        grey_img = utils.color_2_gray(self.original_img)
        inverted_grey = utils.invert_img(grey_img)
        binInv = utils.sperate_fgd(inverted_grey)
        return binInv
    
class Mask:
    def __init__(self, image:cv.Mat):
        self.base_img = image
        self.horizontal = image.copy()
        self.vertical = image.copy()

        self.horizontalSize = image.shape[1]//constants.SCALE # shape[1]->columns
        # structure for horizontal traversal
        self.horizontal_structure = cv.getStructuringElement(cv.MORPH_RECT, (self.horizontalSize, 1))
        # A foreground pixel in the input image will be kept only if all pixels inside the structuring element are > 0. Otherwise, the pixels are set to 0 (i.e., background). 
        self.horizontal_eroded = cv.erode(self.horizontal, self.horizontal_structure, anchor=(-1, -1)) # anchor at the centre of horizontal strcture
        # Perform dilation operation
        self.horizontal_dialated = cv.dilate(self.horizontal_eroded, self.horizontal_structure, anchor=(-1, -1))

        self.verticalSize = image.shape[1]//constants.SCALE # shape[1]->columns
        # Create vertical structuring element
        self.vertical_structure = cv.getStructuringElement(cv.MORPH_RECT, (1, self.verticalSize))
        self.vertical_eroded = cv.erode(self.vertical, self.vertical_structure, anchor=(-1, -1))
        self.vertical_dialated = cv.dilate(self.vertical_eroded, self.vertical_structure, anchor=(-1, -1))

        # Save image for further processing
        # cv2.IMWRITE_PNG_COMPRESSION with a possible value between 0 to 9, the default being 3. 
        # The higher value does high compression of the image resulting in a smaller file size but a longer compression time.
        compression_Params = [cv.IMWRITE_PNG_COMPRESSION, 9]
        cv.imwrite("sliced_images/Horizontal.png", self.horizontal_dialated, compression_Params)
        cv.imwrite("sliced_images/Vertical.png", self.vertical_dialated, compression_Params)

        mask = cv.add(self.horizontal_dialated, self.vertical_dialated)
        mask_points = cv.bitwise_and(self.horizontal_dialated, self.vertical_dialated)
        cv.imwrite('Masked_imgs/mask.png', mask, compression_Params)
        cv.imwrite('Masked_imgs/mask_points.png', mask_points, compression_Params)

    def mean(self, image, x, y, value):
        # Calculate the mean of target region.
        # if (val == 255) => white
        # if (val == 0) => black
        # x =  reference to initial x value (top left corner of pool window)
        # y =  reference to intial y value (top left corner of pool window)
        # r, c - (row, column) points inside pool
        r = x
        c = y
        X = 0
        Y = 0
        count = 0

        while((r- x)<constants.POOL_SIZE and r < image.shape[0]):
            c = y
            while((c-y)<constants.POOL_SIZE and c < image.shape[1]):
                # try:
                #  "(r,c) returns an 3 sized array of RGB values,
                #  since image is grayscaled and binary, all 3 values will be same
                #  "val" is the value that we are looking for -- Black = 0;
                #                                             -- White = 255;
                # 
                # adding all the matched pixels coordinates relative to inital coordinates
                # print(image)
                # print("value is", value)
                # print(r,c)
                # print(image[int(r),int(c)][0])
                
                
                if(image[int(r),int(c)][0] == value):
                    X += r-x
                    Y += c-y
                    count+=1
                c+=1
                # except:
                    # return image, r, c
            r+=1

        if(count!=0):
            X = X//count+x
            Y = Y//count+y
        else:
            X = x+constants.POOL_SIZE//2
            Y = y+constants.POOL_SIZE//2
        print("X, Y :", X, Y)
        return X, Y
    
    def calibrate(self, img, x_in : int, y_in : int, value : int):
        
        X, Y = self.mean(img, x_in, y_in, value)
        
        # img, r, c = mean(img, x_in, y_in, value)
        # compute distance from center to target region mean

        # return img, r, c
        
        distsq = ((x_in + constants.POOL_SIZE // 2) - X) ** 2 + ((y_in + constants.POOL_SIZE // 2 )- Y) ** 2
        # print('distq', distsq)
        if (distsq <= 2): 
            # print("returning")
            return  # if distance is less than or equal to square_root(2),
        # the pool window has been calibrated

        # else move left/right and/or up/down to calibrate so that target region mean is as close as possible to center
        # if (abs(x - (x_in + constants.POOL_SIZE / 2)) > 0):
        #     # *y += ((P.y - (*y + constants.POOL_SIZE / 2)) > 0 ? 1 : -1);
        #     if (x-(x_in+constants.POOL_SIZE /2))>0: # x is the right of the mid point
        #         x_in +=1
        #     else: # x is the left of the middle point
        #         x_in -=1
            
        # if (abs(y - (y_in + constants.POOL_SIZE / 2)) > 0):
        #     if (y-(y_in + constants.POOL_SIZE / 2))>0:
        #         y_in += 1
        #     else:
        #         y_in -= 1

        # Move left/right and/or up/down to calibrate so that the target region mean is as close as possible to the center
        # x_new, y_new = x_in, y_in
        
        if abs(X - (x_in +( constants.POOL_SIZE // 2))) > 0:
            if X - (x_in + constants.POOL_SIZE // 2) > 0:
                x_in += 1 
            else:
                x_in -=1

        if abs(Y - (y_in + constants.POOL_SIZE // 2)) > 0:
            if Y - (y_in + constants.POOL_SIZE // 2) > 0: 
                y_in += 1
            else:
                y_in -=1

        # if abs(x - (x_in + constants.POOL_SIZE // 2)) > 0:
        #     x_new += 1 if x - (x_in + constants.POOL_SIZE // 2) > 0 else -1
        # if abs(y - (y_in + constants.POOL_SIZE // 2)) > 0:
        #     y_new += 1 if y - (y_in + constants.POOL_SIZE // 2) > 0 else -1

        # if x_new == x_in and y_new == y_in:
        #     return x_in, y_in
        # if abs(x - (x + constants.POOL_SIZE // 2)) > 0:
        #     if x - (x + constants.POOL_SIZE // 2) > 0:
        #         x += 1 
        #     else:
        #         x = x-1

        # if abs(y - (y + constants.POOL_SIZE // 2)) > 0:
        #     if y - (y + constants.POOL_SIZE // 2) > 0: 
        #         y += 1
        #     else:
        #         y = y-1
        # else:
        return self.calibrate(img, x_in, y_in, value)
    
    def compare_line(self, A, B):    
        if abs(A.start.y-B.start.y) < constants.POOL_SIZE  * constants.BUFFER_CAPACITY / 2:
            return (A.start.x < B.start.x)
        else:
            return (A.start.y < B.start.y)
    
    def lines(self, img, list_of_lines):
        # This Flag checks a line needed to considered or Not
        line = False

        rows = img.shape[0]
        cols = img.shape[1]

        temp = None
        single_point  = []
        temp_line = [single_point[:] for _ in  range(2)]
        
        # Traversing vertically
        for j in range(0, cols, constants.POOL_SIZE): # y --> sidewards
            buffer = 0 # number
            for i in range(0, rows, constants.POOL_SIZE): # x-> downwards
                # calibrating the current pool window
                # calibrate(img, i, j, 0 if line else 255)
                # img, r, c = 
                # print("caliberating")
                self.calibrate(img, i, j, 0 if line else 255)

                # return img,r,c

                r = i
                print("reseting pool_count")
                pool_count = 0 #stores the count of white of pixels in the pool

                # traversing pool and updating pool count

                while (r-i < constants.POOL_SIZE and r < rows):
                    c =j
                    while(c-j<constants.POOL_SIZE and c<cols):
                        pool_count += int(img[r][c][0] == 255)
                        if img[r][c][0] == 255:
                            print(f"white detection on pool at r, c, bluechannel {r , c, pool_count}")
                        c+=1
                    r+=1
                # print(f"{pool_count}  white regions detected from row_start : {i} and col_start : {j} to row_end : {r} and col_end : {c} ")

                # for buffer to increment, pool_count should be greater than threshold (=constants.POOL_SIZE (considering 1 pixel wide line as minimum) )
                # for white region, bool line will be false for white region as we do not have line in our traversal
                print(f"when line is {line}: line ^ (pool_count >= constants.POOL_SIZE) is: ", line ^ (pool_count >= constants.POOL_SIZE))
                if (line ^ (pool_count >= constants.POOL_SIZE)): # skipping this condition means line continuous/ line is True and count(255) >5
                                                    # Inside this condition when line is false  and count(255) > 5
                                                    # Inside this condition when line is True and count(255) < 5
                    print("in_if_line_is", line)
                    print("line ^ (pool_count >= constants.POOL_SIZE) is: ", line ^ (pool_count >= constants.POOL_SIZE))
                    buffer+=1
                    print("Time to increase buffer:", buffer)

                # storing the first point where buffer started to fill in
                if (buffer == 1): 
                    print("buffer is only 1", i, j)
                    temp = (i, j)
                # boundary condition at image end
                elif (buffer == 0 and (i <= rows - 1 and i >= rows - 1 - constants.POOL_SIZE and line)): 
                    print("bountry condition at the image forced endpoint", i, j)
                    print("buffer is zero means;; white regions detected is less than constants.POOL_SIZE")
                    temp = (i, j)
                # if buffer crosses it's capacity, congrats, you have found the starting/ending point as per requirement
                if ((buffer >= constants.BUFFER_CAPACITY) or (i <= rows - 1 and i >= rows - 1 - constants.POOL_SIZE and line)):
                    print("buffer is greater than buffer capacity : ", constants.BUFFER_CAPACITY)
                    print("you have found the starting/ending point as per requirement", i, j)
                    print(f"Buffer is : {buffer}")
                    # flip bool expression
                    line_1 = line
                    line = not line
                    if (buffer >= constants.BUFFER_CAPACITY):
                        print(f"fliping line from '' {line_1} '' to '' {line} '' because buffer is greater than buffer_capacity ")
                    if (i <= rows - 1 and i >= rows - 1 - constants.POOL_SIZE and line):
                        print("rows is :", rows)
                        print(f"fliping line : '{line}' : because row_i : {i} is <= rows-1 : {rows-1} and row_i >= rows-1-POOL_SIZE : {rows-1-POOL_SIZE} and line is : {line}")

                    buffer = 0
                    if line:

                        temp_line[0] = temp
                        print(temp_line[0], "***********")
                    else:
                        print("other condition", i, j)
                        temp_line[1] = temp
                        print(temp_line[1], "***********")

                        # else if line is false, line has just ended
                        # give the value of end off temp_line, and push the line computed in the array
                        list_of_lines.append(temp_line)
                        temp_line = [single_point[:] for _ in  range(2)]
                        print(list_of_lines, "00000000")
        return list_of_lines
    
    def extract(self):

        horizontal = cv.imread("sliced_images/Horizontal.png")
        vertical = cv.imread("sliced_images/Vertical.png")
        
        self.Vlines = []
        self.Hlines = []

        Filtered_Vlines =[]
        Filtered_Hlines = []

        # transposed_horizontal = cv.transpose(horizontal) # transpose the image so as to use the same functions as used for vertical lines
        transposed_horizontal = cv.transpose(horizontal)
        # The main function, where lines are detected
        # again, does for only vertical lines

        # img, r, c = 
        print("verticaaal")
        self.Vlines = self.lines(vertical, self.Vlines)
        
        # return img,r,c

        print("transposed_horizontal")
        self.Hlines = self.lines(transposed_horizontal, self.Hlines)

        self.Filtered_Hlines = self.sort_coordinates(self.Hlines) 
        self.Filtered_Vlines = self.sort_coordinates(self.Vlines)
        return self.Filtered_Vlines, self.Filtered_Hlines
    
    def draw_line(self, image, Hlines, Vlines):
        img = self.base_img
        for line in Hlines:
            x1, y1 = line[0]
            x2, y2 = line[1]
            # Draw the line
            cv.line(image, (x1, y1), (x2, y2), (0, 0, 255), 2)
        for line in Vlines:
            x1, y1 = line[0]
            x2, y2 = line[1]
            # Draw the line
            cv.line(image, (x1, y1), (x2, y2), (0, 0, 255), 2)
        utils.show(image)
    
    def sort_coordinates(self, coords_list):
        # Ensure there are exactly two coordinates
        sorted_cord = []
        for coords in coords_list:
            if len(coords) != 2:
                raise ValueError("There must be exactly two coordinates")

            # Identify top-left and bottom-right
            top_left = min(coords, key=lambda p: (p[1], p[0]))  # Min by y first, then x
            bottom_right = max(coords, key=lambda p: (p[1], p[0]))  # Max by y first, then x
            sorted_cord.append([top_left, bottom_right])
        return sorted_cord
    
    @staticmethod
    def find_intersection(rectangles):
        if rectangles == []:
            print("no lines found")
            return
        # Initialize the intersection rectangle with the first rectangle
        x1_int, y1_int = rectangles[0][0]
        x2_int, y2_int = rectangles[0][1]

        # Compute intersection with all rectangles
        for rect in rectangles[1:]:
            x1, y1 = rect[0]
            x2, y2 = rect[1]

            # Update the intersection coordinates
            x1_int = min(x1_int, x1)
            y1_int = min(y1_int, y1)
            x2_int = max(x2_int, x2)
            y2_int = max(y2_int, y2)

            # If the rectangles do not intersect, return an empty list or None
            # if x1_int > x2_int or y1_int > y2_int:
            #     return None

        return [(x1_int, y1_int), (x2_int, y2_int)]
    
    def crop_img(self, image, Hlines, Vlines):

        # Loop through each set of coordinates
        for coord in Hlines:
            # Unpack the coordinates
            (x1, y1), (x2, y2) = coord
            
            # Calculate the bounding box
            # Assuming y1 == y2 (horizontal line), you may want to give a small height for cropping
            y_min = y1 - 5  # Adjust the height as needed
            y_max = y1 + 5  # Adjust the height as needed
            x_min = min(x1, x2)
            x_max = max(x1, x2)
            
            # Crop the image
            cropped = image[y_min:y_max, x_min:x_max]
        utils.show(cropped)


class Word:
    table_word_pos = [["word", "start","end"]]
    def __init__(self, text = None, x1 = None, y1 = None, x2=None, y2 = None):
        self.text = text
        self.x1 = x1
        self.y1 = y1
        self.x2 = x2
        self.y2 = y2
        # Define a scaling factor to map pixels to Excel cell positions
        self.scale_x = 1/15 # scaling factor for x-axis
        self.scale_y = 1/15 # scaling factor for y-axis
    def __repr__(self):
        return f"Word(text='{self.text}', x1={self.x1}, y1={self.y1}, x2={self.x2}, y2={self.y2})"
    
    @classmethod
    def read_image(cls, image):
        # Read image
        # image = cv.imread(filename)

        # Initialize empty list for words
        words = []

        # Tesseract OCR
        custom_config = r'--oem 3 --psm 6'  # PSM 6: detect single blocks of words
        text_data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT, config=custom_config)

        # Iterate over each detected word
        for i in range(len(text_data['text'])):
            text = text_data['text'][i]
            conf = int(text_data['conf'][i])
            x, y, w, h = int(text_data['left'][i]), int(text_data['top'][i]), int(text_data['width'][i]), int(text_data['height'][i])
            x2, y2 = x + w, y + h

            # Emplace words only if confidence crosses the threshold
            if conf > constants.PERCENT_CONFIDENCE_THRESHOLD:
                words.append(cls(text, x, y, x2, y2))


        # Return a list of Word objects
        return words
    
    @staticmethod  
    def show_words(img, words):
        for word in words:
            cv.rectangle(img, (word.x1, word.y1), (word.x2, word.y2), (0, 255, 0), 2)
        utils.show(img)
    
    # @staticmethod
    def show_table_words(self, intersection, image, words):
        if intersection == None:
            print("No table to show words")
            return
        for word in words: 
            if (word.x1 >= intersection[0][0] and word.y1 >= intersection[0][1] and
                word.x2 <= intersection[1][0] and word.y2 <= intersection[1][1]):
                print(f"Word '{word.text}' ->({word.x1, word.y1})({word.x2, word.y2}) is inside rectangle.")
                Word.table_word_pos.append([word.text, (word.x1, word.y1), (word.x2, word.y2)])
                color = (0, 255, 0)  # Green if inside
                cv.rectangle(image, (word.x1, word.y1), (word.x2, word.y2), color, 2)
        # print('listwords, ', self.table_word_pos)
        utils.show(image)
    
    def get_excel_position(self, x, y):
        col = get_column_letter(int(x * self.scale_x) + 1)  # Convert x to Excel column letter
        row = int(y * self.scale_y) + 1  # Convert y to Excel row number
        return col, row
    
    def to_excel(self):
        # Create a new Excel workbook and select the active worksheet
        print(Word.table_word_pos)
        wb = Workbook()
        ws = wb.active
        for word_info in Word.table_word_pos[1:]:
            term = word_info[0]
            start_pos = word_info[1]
            print(start_pos,"startpos")
            print(start_pos[0], start_pos[1])
            col, row = self.get_excel_position(start_pos[0], start_pos[1])
            # print(f"{col}{row}", term)
            ws[f"{col}{row}"] = term
        
        # Save the Excel file
        Word.remove_empty_columns(ws)
        Word.remove_empty_rows_before_first(ws)
        wb.save("words_positions.xlsx")

    @staticmethod
    def remove_empty_columns(sheet):
        for col in reversed(range(1, sheet.max_column + 1)):
            col_letter = get_column_letter(col)
            if all(sheet[f"{col_letter}{row}"].value is None for row in range(1, sheet.max_row + 1)):
                sheet.delete_cols(col)
    @staticmethod
    # Function to remove empty rows before the first non-empty row
    def remove_empty_rows_before_first(sheet):
        first_non_empty_row = None
        for row in range(1, sheet.max_row + 1):
            if any(sheet[f"{get_column_letter(col)}{row}"].value for col in range(1, sheet.max_column + 1)):
                first_non_empty_row = row
                break
        if first_non_empty_row:
            for row in range(1, first_non_empty_row):
                sheet.delete_rows(1)

            # return None
    
class Point:
    def __init__(self, x=0, y=0):
        self.x = x
        self.y = y

    def __repr__(self):
        return f"Point({self.x}, {self.y})"
    
class Line:
    def __init__(self, start=None, end=None):
        self.start = start
        self.end = end



# class Word:
#     _instance = None

#     def __new__(cls, *args, **kwargs):
#         if cls._instance is None:
#             cls._instance = super(Word, cls).__new__(cls)
#             cls._instance.table_word_pos = [["word", "start", "end"]]
#         return cls._instance

#     def __init__(self, text=None, x1=None, y1=None, x2=None, y2=None):
#         self.text = text
#         self.x1 = x1
#         self.y1 = y1
#         self.x2 = x2
#         self.y2 = y2
